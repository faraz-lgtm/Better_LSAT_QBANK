#!/usr/bin/env python3
"""Generate PrepTest explanation import templates (CSV + XLSX) from api_json."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
API_JSON_DIR = ROOT / "api_json"
OUT_DIR = ROOT / "data"

EXPLANATION_ROW_COUNT = 500

# Only full prep-test JSON files (LSAC101.json) for the author-facing template.
FULL_PT_JSON = re.compile(r"^LSAC\d+$", re.IGNORECASE)


@dataclass(frozen=True)
class LookupRow:
    module_id: str
    section_id: str
    source_item_id: str
    prep_test: str
    section_number: int
    section_label: str
    question_number: int
    question_label: str
    question_ref: str
    composite_key: str


def prep_test_label(module_id: str) -> str:
    m = re.match(r"^LSAC(\d+)$", module_id.strip(), re.IGNORECASE)
    if m:
        return f"PT {int(m.group(1))}"
    return module_id.strip()


def section_label_for(number: int) -> str:
    return f"S{number}"


def question_label_for(number: int) -> str:
    return f"Q{number}"


def infer_section_code(section_name: str | None) -> str:
    if not section_name:
        return ""
    value = section_name.lower()
    if "logical reasoning" in value or value == "lr":
        return "LR"
    if "reading comprehension" in value or value == "rc":
        return "RC"
    if "logic game" in value or "analytical reasoning" in value or value == "lg":
        return "LG"
    head = (section_name or "")[:2].upper()
    if head in ("LR", "RC", "LG"):
        return head
    return ""


SAMPLE_EXPLANATIONS: list[tuple[str, str, str, str]] = [
    (
        "PT 101",
        "S1",
        "Q1",
        "<p><strong>Correct answer: B</strong></p>"
        "<p>Example HTML explanation — replace with your content.</p>",
    ),
]


def load_all_lookup_rows() -> list[LookupRow]:
    rows: list[LookupRow] = []
    seen: set[tuple[str, str, str]] = set()

    for path in sorted(API_JSON_DIR.glob("*.json")):
        if not FULL_PT_JSON.match(path.stem):
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        sections = data.get("sections")
        if not isinstance(sections, list):
            continue

        module_id = path.stem
        prep_test = prep_test_label(module_id)

        for section in sections:
            if not isinstance(section, dict):
                continue
            section_id = str(section.get("sectionId", "")).strip()
            if not section_id:
                continue
            section_number = section.get("sectionOrder")
            if not isinstance(section_number, int):
                section_number = 1
            sec_label = section_label_for(section_number)
            section_name = str(section.get("sectionName") or "")
            section_code = infer_section_code(section_name)

            items = section.get("items")
            if not isinstance(items, list):
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                item_id = str(item.get("itemId", "")).strip()
                if not item_id:
                    continue
                qnum = item.get("itemPosition")
                if not isinstance(qnum, int):
                    continue
                dedupe = (module_id, section_id, item_id)
                if dedupe in seen:
                    continue
                seen.add(dedupe)

                qlabel = question_label_for(qnum)
                qref = f"{prep_test} · {sec_label} · {qlabel}"
                if section_code:
                    qref = f"{prep_test} · {sec_label} ({section_code}) · {qlabel}"
                composite = f"{prep_test}|{sec_label}|{qlabel}"

                rows.append(
                    LookupRow(
                        module_id=module_id,
                        section_id=section_id,
                        source_item_id=item_id,
                        prep_test=prep_test,
                        section_number=section_number,
                        section_label=sec_label,
                        question_number=qnum,
                        question_label=qlabel,
                        question_ref=qref,
                        composite_key=composite,
                    )
                )

    rows.sort(key=lambda r: (r.prep_test, r.section_number, r.question_number))
    return rows


def unique_prep_tests(rows: list[LookupRow]) -> list[str]:
    tests: list[str] = []
    prev = None
    for row in rows:
        if row.prep_test != prev:
            tests.append(row.prep_test)
            prev = row.prep_test
    return tests


def write_csv_lookup(rows: list[LookupRow], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "question_ref",
                "prep_test",
                "section_number",
                "section_label",
                "question_number",
                "question_label",
                "module_id",
                "section_id",
                "source_item_id",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.question_ref,
                    row.prep_test,
                    row.section_number,
                    row.section_label,
                    row.question_number,
                    row.question_label,
                    row.module_id,
                    row.section_id,
                    row.source_item_id,
                ]
            )


def write_csv_explanations(path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["prep_test", "section", "question", "explanation"])
        writer.writerows(SAMPLE_EXPLANATIONS)


def write_xlsx(rows: list[LookupRow], prep_tests: list[str], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required for XLSX output. Install with: pip install openpyxl"
        ) from e

    last_lookup = len(rows) + 1
    last_expl_row = 1 + EXPLANATION_ROW_COUNT
    header_fill = PatternFill("solid", fgColor="E8E8E8")

    wb = Workbook()

    readme = wb.active
    readme.title = "README"
    readme["A1"] = "PrepTest explanations — author template"
    readme["A2"] = (
        "On Explanations: choose Prep Test (e.g. PT 101), Section (S1–S4), "
        "Question (Q1–Q27), then enter HTML in column D."
    )
    readme["A3"] = (
        f"Covers {len(prep_tests)} full prep tests ({len(rows)} questions) from LSAC###.json. "
        "Lookup sheet has internal IDs for import; do not reorder Lookup rows."
    )
    readme["A4"] = 'Reference format matches the app: "PT 92 · S2 · Q4".'
    readme["A1"].font = Font(bold=True, size=12)
    for r in (2, 3, 4):
        readme[f"A{r}"].alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 95

    lookup_ws = wb.create_sheet("Lookup")
    lookup_headers = [
        "question_ref",
        "prep_test",
        "section_label",
        "question_label",
        "section_number",
        "question_number",
        "module_id",
        "section_id",
        "source_item_id",
        "_key",
    ]
    lookup_ws.append(lookup_headers)
    for row in rows:
        lookup_ws.append(
            [
                row.question_ref,
                row.prep_test,
                row.section_label,
                row.question_label,
                row.section_number,
                row.question_number,
                row.module_id,
                row.section_id,
                row.source_item_id,
                row.composite_key,
            ]
        )
    for cell in lookup_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    lookup_ws.column_dimensions["A"].width = 36
    lookup_ws.column_dimensions["B"].width = 12
    lookup_ws.column_dimensions["C"].width = 10
    lookup_ws.column_dimensions["D"].width = 10
    for col in ("E", "F", "G", "H", "I", "J"):
        lookup_ws.column_dimensions[col].hidden = True

    prep_tests_ws = wb.create_sheet("PrepTests")
    prep_tests_ws["A1"] = "prep_test"
    prep_tests_ws["A1"].font = Font(bold=True)
    prep_tests_ws["A1"].fill = header_fill
    for idx, label in enumerate(prep_tests, start=2):
        prep_tests_ws.cell(row=idx, column=1, value=label)
    prep_tests_ws.column_dimensions["A"].width = 14
    last_prep_test = len(prep_tests) + 1

    expl_ws = wb.create_sheet("Explanations")
    expl_ws.append(["prep_test", "section", "question", "explanation"])
    for sample in SAMPLE_EXPLANATIONS:
        expl_ws.append(list(sample))
    for _ in range(EXPLANATION_ROW_COUNT - len(SAMPLE_EXPLANATIONS)):
        expl_ws.append(["", "", "", ""])
    for cell in expl_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    expl_ws.column_dimensions["A"].width = 14
    expl_ws.column_dimensions["B"].width = 10
    expl_ws.column_dimensions["C"].width = 10
    expl_ws.column_dimensions["D"].width = 72
    for r in range(2, last_expl_row + 1):
        expl_ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    dv_prep = DataValidation(
        type="list",
        formula1=f"=PrepTests!$A$2:$A${last_prep_test}",
        allow_blank=True,
    )
    expl_ws.add_data_validation(dv_prep)
    dv_prep.add(f"A2:A{last_expl_row}")

    dv_section = DataValidation(
        type="list",
        formula1=(
            f"=OFFSET(Lookup!$C$1,MATCH(A2,Lookup!$B$2:$B${last_lookup},0),"
            f"0,COUNTIF(Lookup!$B$2:$B${last_lookup},A2),1)"
        ),
        allow_blank=True,
    )
    expl_ws.add_data_validation(dv_section)
    dv_section.add(f"B2:B{last_expl_row}")

    dv_question = DataValidation(
        type="list",
        formula1=(
            f'=OFFSET(Lookup!$D$1,MATCH(A2&"|"&B2,Lookup!$J$2:$J${last_lookup},0),'
            f"0,COUNTIFS(Lookup!$B$2:$B${last_lookup},A2,"
            f"Lookup!$C$2:$C${last_lookup},B2),1)"
        ),
        allow_blank=True,
    )
    expl_ws.add_data_validation(dv_question)
    dv_question.add(f"C2:C{last_expl_row}")

    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    lookup_rows = load_all_lookup_rows()
    if not lookup_rows:
        raise SystemExit(f"No lookup rows found under {API_JSON_DIR}")

    prep_tests = unique_prep_tests(lookup_rows)
    lookup_csv = OUT_DIR / "preptest_explanations_lookup.csv"
    sample_csv = OUT_DIR / "preptest_explanations_sample.csv"
    xlsx_path = OUT_DIR / "preptest_explanations_import_template.xlsx"

    write_csv_lookup(lookup_rows, lookup_csv)
    write_csv_explanations(sample_csv)
    write_xlsx(lookup_rows, prep_tests, xlsx_path)

    print(f"Wrote {lookup_csv} ({len(lookup_rows)} questions, {len(prep_tests)} prep tests)")
    print(f"Wrote {sample_csv}")
    print(f"Wrote {xlsx_path} ({EXPLANATION_ROW_COUNT} editable rows)")


if __name__ == "__main__":
    main()
